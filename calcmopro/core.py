"""Business logic for CALCMO Pro.

This module keeps the calculation pipeline independent from the desktop UI:
read the source workbook, compute MGA/MO results, generate a styled workbook,
and create an input template.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, GradientFill, PatternFill, Side
from openpyxl.utils import get_column_letter


MATIERES = [
    {"id": "redaction", "label": "Redaction", "coef": 2, "anglais": False},
    {"id": "maths", "label": "Mathematiques", "coef": 2, "anglais": False},
    {"id": "pc", "label": "P.C.", "coef": 1, "anglais": False},
    {"id": "anglais", "label": "Anglais", "coef": 1, "anglais": True},
]

TOTAL_COEF = sum(matiere["coef"] for matiere in MATIERES)

MENTIONS = [
    (10, "Felicitations !"),
    (0, "Insuffisant"),
]

REQUIRED_COLUMNS = [
    "NOM",
    "CLASSE",
    "ETABLISSEMENT",
    "REDACTION_T1",
    "REDACTION_T2",
    "REDACTION_T3",
    "MATHS_T1",
    "MATHS_T2",
    "MATHS_T3",
    "PC_T1",
    "PC_T2",
    "PC_T3",
    "ANGLAIS_T1",
    "ANGLAIS_T2",
    "ANGLAIS_T3",
    "REDACTION_BEPC",
    "MATHS_BEPC",
    "PC_BEPC",
    "ANGLAIS_ECRIT",
    "ANGLAIS_ORAL",
]

# Workbook palette.
C_NAVY = "0A1628"
C_BLUE = "1E4D9B"
C_CYAN = "38BDF8"
C_GOLD = "FBBF24"
C_GREEN = "22C55E"
C_RED = "EF4444"
C_ORANGE = "F97316"
C_WHITE = "F8FAFC"
C_MUTED = "94A3B8"
C_HEADER = "0F2044"
C_ROW1 = "0D1F3C"
C_ROW2 = "0A1628"
C_LINE = "1A3A6E"


@dataclass(frozen=True)
class ImportReport:
    students: list[dict[str, Any]]
    warnings: list[str]
    source_columns: list[str]


def calc_mga(t1: float | None, t2: float | None, t3: float | None) -> float | None:
    """MGA = (T1 + 2*T2 + 2*T3) / available weights."""
    values = {"t1": t1, "t2": t2, "t3": t3}
    weights = {"t1": 1, "t2": 2, "t3": 2}
    numerator = sum(values[key] * weights[key] for key in values if values[key] is not None)
    denominator = sum(weights[key] for key in values if values[key] is not None)
    if denominator == 0:
        return None
    return numerator / denominator


def calc_note_bepc_anglais(ecrit: float | None, oral: float | None) -> float | None:
    """English BEPC mark = (written + oral) / 2."""
    if ecrit is None or oral is None:
        return None
    return (ecrit + oral) / 2


def get_mention(mo: float) -> str:
    return next(mention for threshold, mention in MENTIONS if mo >= threshold)


def calc_mo(eleve: dict[str, Any]) -> dict[str, Any]:
    """Compute MO result rows for one student."""
    total = 0.0
    lignes: list[dict[str, Any]] = []

    for matiere in MATIERES:
        mid = matiere["id"]
        coef = matiere["coef"]
        data = eleve.get("matieres", {}).get(mid, {})
        mga = data.get("mga")

        if matiere["anglais"]:
            note_bepc = calc_note_bepc_anglais(data.get("bepc_ecrit"), data.get("bepc_oral"))
        else:
            note_bepc = data.get("bepc")

        if mga is None or note_bepc is None:
            somme = None
            produit = None
        else:
            somme = mga + note_bepc
            produit = somme * coef
            total += produit

        lignes.append(
            {
                "matiere": matiere["label"],
                "coef": coef,
                "mga": mga,
                "note_bepc": note_bepc,
                "somme": somme,
                "produit": produit,
            }
        )

    mo = total / (2 * TOTAL_COEF)
    return {
        "nom": eleve.get("nom", ""),
        "classe": eleve.get("classe", ""),
        "etablissement": eleve.get("etablissement", ""),
        "lignes": lignes,
        "total": total,
        "mo": mo,
        "mention": get_mention(mo),
    }


def compute_results(students: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [calc_mo(student) for student in students]


def summarize_results(resultats: list[dict[str, Any]]) -> dict[str, Any]:
    mos = [resultat["mo"] for resultat in resultats]
    total = len(mos)
    admitted = sum(1 for mo in mos if mo >= 10)
    insufficient = total - admitted
    mention_counts = {mention: 0 for _, mention in MENTIONS}
    for resultat in resultats:
        mention_counts[resultat["mention"]] = mention_counts.get(resultat["mention"], 0) + 1

    return {
        "effectif": total,
        "mo_min": min(mos) if mos else None,
        "mo_max": max(mos) if mos else None,
        "mo_moyenne": sum(mos) / total if total else None,
        "admis": admitted,
        "insuffisants": insufficient,
        "taux_admis": admitted / total if total else 0,
        "mentions": mention_counts,
    }


def read_input_file(path: str | Path) -> ImportReport:
    """Read an Excel input file and return normalized students plus warnings."""
    input_path = Path(path)
    if not input_path.exists():
        raise FileNotFoundError(f"Fichier introuvable : {input_path}")
    if input_path.suffix.lower() not in {".xlsx", ".xlsm", ".xls"}:
        raise ValueError("Le fichier d'entree doit etre un classeur Excel.")

    df = pd.read_excel(input_path, dtype=str)
    df.columns = [str(column).strip().upper() for column in df.columns]
    missing_columns = [column for column in REQUIRED_COLUMNS if column not in df.columns]

    warnings: list[str] = []
    if missing_columns:
        warnings.append("Colonnes manquantes : " + ", ".join(missing_columns))

    def num(row: pd.Series, column: str) -> float | None:
        value = row.get(column)
        if pd.isna(value) or value in (None, "", "NAN", "nan"):
            return None
        try:
            number = float(str(value).strip().replace(",", "."))
        except ValueError:
            warnings.append(f"Valeur ignoree dans {column} : {value}")
            return None
        if number < 0 or number > 20:
            warnings.append(f"Note hors plage 0-20 dans {column} : {value}")
        return number

    students: list[dict[str, Any]] = []
    for index, row in df.iterrows():
        nom = str(row.get("NOM", "")).strip()
        if not nom or nom.upper() == "NAN":
            continue

        def mga(prefix: str) -> float | None:
            return calc_mga(
                num(row, f"{prefix}_T1"),
                num(row, f"{prefix}_T2"),
                num(row, f"{prefix}_T3"),
            )

        student = {
            "nom": nom,
            "classe": _clean_text(row.get("CLASSE", "")),
            "etablissement": _clean_text(row.get("ETABLISSEMENT", "")),
            "matieres": {
                "redaction": {
                    "mga": mga("REDACTION"),
                    "bepc": num(row, "REDACTION_BEPC"),
                },
                "maths": {
                    "mga": mga("MATHS"),
                    "bepc": num(row, "MATHS_BEPC"),
                },
                "pc": {
                    "mga": mga("PC"),
                    "bepc": num(row, "PC_BEPC"),
                },
                "anglais": {
                    "mga": mga("ANGLAIS"),
                    "bepc_ecrit": num(row, "ANGLAIS_ECRIT"),
                    "bepc_oral": num(row, "ANGLAIS_ORAL"),
                },
            },
        }
        students.append(student)

        if all(line.get("mga") is None for line in student["matieres"].values()):
            warnings.append(f"Ligne {index + 2} : aucune moyenne trimestrielle exploitable.")

    if not students:
        warnings.append("Aucun eleve avec un NOM valide n'a ete trouve.")

    return ImportReport(students=students, warnings=deduplicate(warnings), source_columns=list(df.columns))


def default_output_path(input_path: str | Path) -> Path:
    path = Path(input_path)
    return path.with_name(path.stem + "_RESULTATS.xlsx")


def generate_template(output_path: str | Path) -> Path:
    """Create a professional Excel input template."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SAISIE"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    for column_index, header in enumerate(REQUIRED_COLUMNS, start=1):
        cell = ws.cell(row=1, column=column_index, value=header)
        cell.font = _font(bold=True, size=9, color=C_WHITE)
        cell.fill = _fill(C_BLUE)
        cell.alignment = _align(wrap=True)
        cell.border = _border()
        ws.column_dimensions[get_column_letter(column_index)].width = 16

    samples = [
        [
            "KOUAKOU Jean-Paul",
            "3eme A",
            "LM Abengourou",
            9.5,
            10,
            10.5,
            11,
            11.5,
            11,
            10.5,
            11,
            10.5,
            13.5,
            14,
            15,
            10,
            8,
            10,
            14,
            13,
        ],
        [
            "BROU Ange-Marie",
            "3eme B",
            "LM Abengourou",
            14,
            14.5,
            15,
            16,
            15.5,
            16.5,
            13.5,
            14,
            14,
            15,
            15.5,
            16,
            13,
            15,
            14,
            16,
            15,
        ],
    ]
    for row_index, row_values in enumerate(samples, start=2):
        for column_index, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_index, column=column_index, value=value)
            cell.fill = _fill("F8FAFC" if row_index % 2 == 0 else "EEF2F7")
            cell.alignment = _align()
            cell.border = _border(color="D8E0EA")
            if column_index <= 3:
                cell.alignment = _align(h="left")

    notes = wb.create_sheet("REFERENTIEL")
    notes.sheet_view.showGridLines = False
    notes["A1"] = "Colonnes attendues"
    notes["A1"].font = _font(bold=True, size=13, color=C_GOLD)
    notes["A1"].fill = _fill(C_HEADER)
    notes["A1"].alignment = _align(h="left")
    notes.merge_cells("A1:C1")

    info_rows = [
        ("MGA", "(T1 + 2*T2 + 2*T3) / 5", "Si une note manque, les poids disponibles sont utilises."),
        ("MO", "Somme[(MGA + Note BEPC) * coeff] / (2 * somme coeff)", "Coeff total = 6."),
        ("Anglais", "Note BEPC = (Ecrit + Oral) / 2", "Les deux notes sont necessaires."),
    ]
    for row_index, values in enumerate(info_rows, start=3):
        for column_index, value in enumerate(values, start=1):
            cell = notes.cell(row=row_index, column=column_index, value=value)
            cell.fill = _fill(C_ROW1 if row_index % 2 else C_ROW2)
            cell.font = _font(size=9, color=C_WHITE, bold=column_index == 1)
            cell.alignment = _align(h="left", wrap=True)
            cell.border = _border(color=C_LINE)

    for column_index, width in enumerate([22, 42, 52], start=1):
        notes.column_dimensions[get_column_letter(column_index)].width = width

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def generate_excel(
    resultats: list[dict[str, Any]],
    output_path: str | Path,
    etablissement: str = "",
    annee: str = "2025-2026",
) -> Path:
    """Generate the styled results workbook."""
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "RECAPITULATIF MO"
    _build_recap(ws1, resultats, etablissement, annee)

    ws2 = wb.create_sheet("DETAIL CALCULS")
    _build_detail(ws2, resultats)

    ws3 = wb.create_sheet("STATISTIQUES")
    _build_stats(ws3, resultats)

    wb.save(path)
    return path


def format_number(value: float | None, decimals: int = 3) -> str:
    if value is None:
        return "-"
    return f"{value:.{decimals}f}".replace(".", ",")


def deduplicate(items: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for item in items:
        if item not in seen:
            result.append(item)
            seen.add(item)
    return result


def _clean_text(value: Any) -> str:
    if pd.isna(value) or value is None:
        return ""
    text = str(value).strip()
    return "" if text.upper() == "NAN" else text


def _font(bold: bool = False, size: int = 10, color: str = C_WHITE, name: str = "Calibri") -> Font:
    return Font(bold=bold, size=size, color=color, name=name)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _border(style: str = "thin", color: str = "2D4A7A") -> Border:
    side = Side(style=style, color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _align(h: str = "center", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _mention_color(mention: str) -> str:
    mapping = {
        "Felicitations !": C_GREEN,
        "Insuffisant": C_RED,
    }
    return mapping.get(mention, C_WHITE)


def _build_recap(ws: openpyxl.worksheet.worksheet.Worksheet, resultats: list[dict[str, Any]], etablissement: str, annee: str) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C_CYAN
    ws.freeze_panes = "A4"

    ws.merge_cells("A1:J1")
    title = ws["A1"]
    title.value = "CALCMO PRO - MOYENNE D'ORIENTATION BEPC"
    title.font = _font(bold=True, size=15, color=C_GOLD)
    title.fill = GradientFill(stop=(C_HEADER, C_NAVY))
    title.alignment = _align()

    ws.merge_cells("A2:J2")
    subtitle = ws["A2"]
    subtitle.value = f"{etablissement or 'Etablissement'} | Annee scolaire {annee} | Genere le {date.today().strftime('%d/%m/%Y')}"
    subtitle.font = _font(size=10, color=C_MUTED)
    subtitle.fill = _fill(C_NAVY)
    subtitle.alignment = _align()

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20

    headers = ["N", "NOM ET PRENOMS", "CLASSE", "REDACTION", "MATHS", "P.C.", "ANGLAIS", "TOTAL", "MO", "MENTION"]
    for column, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=column, value=header)
        cell.font = _font(bold=True, size=9, color=C_WHITE)
        cell.fill = _fill(C_BLUE)
        cell.alignment = _align(wrap=True)
        cell.border = _border()
    ws.row_dimensions[3].height = 24

    for index, resultat in enumerate(resultats, start=1):
        row = 3 + index
        background = C_ROW1 if index % 2 else C_ROW2
        values: list[Any] = [index, resultat["nom"], resultat["classe"]]
        for matiere in MATIERES:
            ligne = next((item for item in resultat["lignes"] if item["matiere"] == matiere["label"]), {})
            values.append(ligne.get("mga"))
        values.extend([resultat["total"], resultat["mo"], resultat["mention"]])

        for column, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=column, value=value)
            cell.fill = _fill(background)
            cell.alignment = _align(h="left" if column == 2 else "center")
            cell.border = _border(color=C_LINE)
            if isinstance(value, float):
                cell.number_format = "0.0000" if column == 9 else "0.000"
            if column == 9:
                cell.font = _font(bold=True, size=9, color=C_GREEN if value >= 10 else C_RED)
            elif column == 10:
                cell.font = _font(bold=True, size=9, color=_mention_color(str(value)))
            else:
                cell.font = _font(size=9, color=C_WHITE)

    widths = [6, 34, 12, 13, 13, 11, 12, 13, 12, 16]
    for column, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(column)].width = width

    last_row = max(4, len(resultats) + 3)
    ws.auto_filter.ref = f"A3:J{last_row}"


def _build_detail(ws: openpyxl.worksheet.worksheet.Worksheet, resultats: list[dict[str, Any]]) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C_GOLD
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:K1")
    title = ws["A1"]
    title.value = "DETAIL DES CALCULS PAR ELEVE"
    title.font = _font(bold=True, size=13, color=C_GOLD)
    title.fill = GradientFill(stop=(C_HEADER, C_NAVY))
    title.alignment = _align()
    ws.row_dimensions[1].height = 28

    headers = ["NOM", "CLASSE", "MATIERE", "COEFF.", "MGA", "NOTE BEPC", "MGA+BEPC", "PRODUIT", "TOTAL", "MO", "MENTION"]
    for column, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=column, value=header)
        cell.font = _font(bold=True, size=8, color=C_WHITE)
        cell.fill = _fill(C_BLUE)
        cell.alignment = _align(wrap=True)
        cell.border = _border()
    ws.row_dimensions[2].height = 22

    row = 3
    for resultat in resultats:
        for line_index, line in enumerate(resultat["lignes"]):
            background = C_ROW1 if row % 2 else C_ROW2
            values: list[Any] = [
                resultat["nom"] if line_index == 0 else "",
                resultat["classe"] if line_index == 0 else "",
                line["matiere"],
                line["coef"],
                line["mga"],
                line["note_bepc"],
                line["somme"],
                line["produit"],
                resultat["total"] if line_index == 0 else "",
                resultat["mo"] if line_index == 0 else "",
                resultat["mention"] if line_index == 0 else "",
            ]
            for column, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=column, value=value)
                cell.fill = _fill(background)
                cell.font = _font(size=8, color=C_WHITE, bold=column in (9, 10, 11) and line_index == 0)
                cell.alignment = _align(h="left" if column in (1, 2, 3) else "center")
                cell.border = _border(color=C_LINE)
                if isinstance(value, float):
                    cell.number_format = "0.0000" if column == 10 else "0.000"
            row += 1

        for column in range(1, 12):
            cell = ws.cell(row=row, column=column, value="")
            cell.fill = _fill("071224")
        row += 1

    widths = [32, 12, 18, 8, 10, 12, 12, 12, 12, 12, 16]
    for column, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(column)].width = width


def _build_stats(ws: openpyxl.worksheet.worksheet.Worksheet, resultats: list[dict[str, Any]]) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C_GREEN

    ws.merge_cells("A1:E1")
    title = ws["A1"]
    title.value = "STATISTIQUES - MOYENNE D'ORIENTATION"
    title.font = _font(bold=True, size=13, color=C_GOLD)
    title.fill = GradientFill(stop=(C_HEADER, C_NAVY))
    title.alignment = _align()
    ws.row_dimensions[1].height = 28

    summary = summarize_results(resultats)
    stats = [
        ("Effectif total", summary["effectif"], ""),
        ("MO minimale", summary["mo_min"], ""),
        ("MO maximale", summary["mo_max"], ""),
        ("MO moyenne", summary["mo_moyenne"], ""),
        ("Felicitations (MO >= 10)", summary["admis"], f"{summary['taux_admis'] * 100:.1f}%"),
        ("Insuffisants (MO < 10)", summary["insuffisants"], f"{(1 - summary['taux_admis']) * 100:.1f}%"),
    ]

    for column, header in enumerate(["Indicateur", "Valeur", "%"], start=1):
        cell = ws.cell(row=2, column=column, value=header)
        cell.font = _font(bold=True, size=9, color=C_WHITE)
        cell.fill = _fill(C_BLUE)
        cell.alignment = _align()
        cell.border = _border()

    for index, values in enumerate(stats, start=3):
        background = C_ROW1 if index % 2 else C_ROW2
        for column, value in enumerate(values, start=1):
            cell = ws.cell(row=index, column=column, value=value)
            cell.fill = _fill(background)
            cell.font = _font(size=9, color=C_WHITE, bold=column == 2)
            cell.alignment = _align(h="left" if column == 1 else "center")
            cell.border = _border(color=C_LINE)
            if isinstance(value, float):
                cell.number_format = "0.0000"

    mention_title_row = 11
    ws.merge_cells(start_row=mention_title_row, start_column=1, end_row=mention_title_row, end_column=3)
    mention_title = ws.cell(row=mention_title_row, column=1, value="REPARTITION PAR MENTION")
    mention_title.font = _font(bold=True, size=10, color=C_CYAN)
    mention_title.fill = _fill(C_NAVY)
    mention_title.alignment = _align(h="left")

    header_row = mention_title_row + 1
    for column, header in enumerate(["Mention", "Effectif", "%"], start=1):
        cell = ws.cell(row=header_row, column=column, value=header)
        cell.font = _font(bold=True, size=9, color=C_WHITE)
        cell.fill = _fill(C_BLUE)
        cell.alignment = _align()
        cell.border = _border()

    total = summary["effectif"]
    for index, (_, mention) in enumerate(MENTIONS, start=header_row + 1):
        count = summary["mentions"].get(mention, 0)
        percent = count / total if total else 0
        for column, value in enumerate([mention, count, percent], start=1):
            cell = ws.cell(row=index, column=column, value=value)
            cell.fill = _fill(C_ROW1 if index % 2 else C_ROW2)
            cell.font = _font(size=9, color=_mention_color(mention) if column == 1 else C_WHITE, bold=column == 2)
            cell.alignment = _align(h="left" if column == 1 else "center")
            cell.border = _border(color=C_LINE)
            if column == 3:
                cell.number_format = "0.0%"

    if total:
        chart = BarChart()
        chart.title = "Effectifs par mention"
        chart.y_axis.title = "Effectif"
        chart.x_axis.title = "Mention"
        data = Reference(ws, min_col=2, min_row=header_row, max_row=header_row + len(MENTIONS))
        categories = Reference(ws, min_col=1, min_row=header_row + 1, max_row=header_row + len(MENTIONS))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 7
        chart.width = 12
        ws.add_chart(chart, "E12")

    for column, width in enumerate([30, 14, 12, 3, 18], start=1):
        ws.column_dimensions[get_column_letter(column)].width = width
